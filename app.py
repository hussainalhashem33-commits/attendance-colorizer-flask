
import os
from flask import Flask, render_template, request, send_file
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tempfile

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload size

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_xlsx(file_path):
    # Load the workbook
    wb = load_workbook(file_path)

    # Define colors
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Iterate through cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_value = str(cell.value).strip().lower()
                    if cell_value == 'a':  # Absence
                        cell.fill = red_fill
                    elif cell_value == 'p':  # Presence
                        cell.fill = yellow_fill

    # Save to a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()

    return temp_file.name

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods:['POST'])
def upload_file():
    if 'file' not in request.files:
        return render_template('index.html', error='No file part')

    file = request.files['file']

    if file.filename == '':
        return render_template('index.html', error='No selected file')

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        # Process the file
        processed_file_path = process_xlsx(file_path)

        # Return the processed file for download
        return send_file(
            processed_file_path,
            as_attachment=True,
            download_name=filename,
            mimetype:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return render_template('index.html', error:'Invalid file type. Please upload an XLSX file.')

if __name__ == '__main__':
    app.run(debug=True)
